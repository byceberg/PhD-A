import openpyxl
import pyautogui
import pyperclip
import time
from typing import Optional

class TTS():
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "2016 Verileri"
        
        self.year = (790, 293)
        self.year_2016 = (752, 353)
        self.trash_code = (1190, 373)
        self.city = (775, 329)
        self.gap = (1084, 538)
        self.regenerate = (593, 561)
        self.calculation_start = (488, 556)
        self.calculation_end = (531, 556)
        self.write_res = (540, 583)
        self.search = (480, 650)
        self.error = (853, 570)
        self.okey = (937, 651)
        self.res = (770, 723)
        self.tr_res = (672, 725)
        
        self.cities = ["BATMAN", "MARDIN", "SIIRT", "SIRNAK", "DIYARBAKIR", "SANLIURFA", "ADIYAMAN", "GAZIANTEP", "KILIS", "KAHRAMANMARAS", "TR"]
        self.trash_codes = ["010505", "010506", "050102", "050103", "050105", "050106", "050107", "050108", "050109", "050110", "050116", "050199", "130502", "130503", "130508", "160508", "160804", "160805", "160807", "170503"]
    
    def move_and_click(self, coordinates : Optional[tuple]):
        pyautogui.moveTo(coordinates[0], coordinates[1])
        pyautogui.leftClick()
    
    def drag(self, target : Optional[tuple]):
        pyautogui.mouseDown()
        pyautogui.moveTo(target[0], target[1])
        pyautogui.mouseUp()
    
    def copy(self):
        pyautogui.hotkey("ctrl", "c")
        time.sleep(0.2)
        res = pyperclip.paste()
        
        return res
    
    def get_value(self, r, c, tr_i):
        pyautogui.moveTo(self.res[0], self.res[1]) if tr_i != 10 else pyautogui.moveTo(self.tr_res[0], self.tr_res[1])
        pyautogui.doubleClick()
        
        value = self.copy()
        self.sheet.cell(r, c, value)
    
    def check(self, row, col, tr_i):
        time.sleep(0.5)
        
        pyautogui.moveTo(self.error[0], self.error[1])
        rgb = pyautogui.pixel(self.error[0], self.error[1])
        
        if rgb == (255, 0, 0):
            print("Couldn't Found")
            self.sheet.cell(row, col, "X") 
        
        else:
            print("Found")
            self.get_value(row, col, tr_i)
    
    def main(self):
        row = 1
        
        for code in self.trash_codes:    
            col = 1
            tr_i = 0
            
            self.move_and_click(self.year)
            self.move_and_click(self.year_2016)
            self.move_and_click(self.trash_code)
            pyautogui.doubleClick()
            pyautogui.write(code)
                
            for city in self.cities:    
                self.move_and_click(self.city)
                pyautogui.moveTo(self.city[0], self.city[1] + 30)
                
                for i in range(5):
                    pyautogui.scroll(600)
                    time.sleep(0.001)
                
                self.move_and_click((self.city[0], self.city[1] + 20))
                
                if tr_i < 10:
                    self.move_and_click(self.city)
                    pyautogui.write(city)
                
                self.move_and_click(self.gap)    
                self.move_and_click(self.regenerate)
                self.move_and_click(self.calculation_start)
                self.drag(self.calculation_end)
                res = self.copy()
                res = eval(res)
                self.move_and_click(self.write_res)
                pyautogui.write(str(res))
                self.move_and_click(self.search)
                self.check(row, col, tr_i)
                print(tr_i)
                col += 1
            
                self.workbook.save("Excel_2016.xlsx")
                time.sleep(9)

                tr_i += 1
            
            row += 1

TTS().main()